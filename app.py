"""
상업용 부동산 뉴스 클리핑 v4
- 네이버 뉴스 API 단독 수집
- 링크 리다이렉트 해석 + 본문 크롤링 + 원문 제목/언론사 보정
- Gemini REST API 명사형 요약 (병렬 처리)
- 토큰 블로킹 기반 고속 중복 제거
"""
import io
import re
import time
import html
import inspect as _inspect
import threading
import datetime as dt
import urllib.parse
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import pandas as pd
import streamlit as st

try:
    from newspaper import Article, Config as NewspaperConfig
    NEWSPAPER_AVAILABLE = True
except ImportError:
    NEWSPAPER_AVAILABLE = False

try:
    import trafilatura
except ImportError:
    trafilatura = None

try:
    from bs4 import BeautifulSoup
    BS_AVAILABLE = True
except ImportError:
    BS_AVAILABLE = False

st.set_page_config(page_title="상업용 부동산 뉴스 클리핑", page_icon="📰", layout="wide")

KST = dt.timezone(dt.timedelta(hours=9))
PRESS_PLACEHOLDER = "(언론사 기입 필요)"
MAIL_CATEGORIES = ["개발계획", "매입매각", "이전동향", "업계동향", "시장동향", "정책"]

UA_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
}

DEFAULT_KEYWORDS = {
    "기존 키워드": [
        "자산운용 매각", "자산운용 매입", "복합개발 -분양", "리테일 상권", "물류센터 매매",
        "물류센터 공실", "오피스 이전 -영화", "매각주관사 빌딩", "사옥 매각", "리츠 건물",
        "오피스 복합개발", "부동산 복합개발", "오피스 매입", "사옥 이전", "사옥 신축",
        "사무실 이전", "물류센터 매각", "물류센터 투자", "증권 부동산 투자 -분양",
        "오피스 펀드", "오피스 리츠", "공유 오피스", "물류센터 부동산", "데이터센터 개발",
        "데이터센터 투자", "증권 부동산 투자 해외 -분양", "보험업",
    ],
    "신규 키워드": [],
}


# ══════════════════════════════════════════════════════════════
# 공통 유틸
# ══════════════════════════════════════════════════════════════
def get_secret(name: str, default: str = "") -> str:
    """secrets.toml이 없는 환경에서도 죽지 않는 안전한 secrets 접근."""
    try:
        return st.secrets.get(name, default)
    except Exception:
        return default


def _st_version():
    try:
        return tuple(int(x) for x in re.findall(r"\d+", st.__version__)[:2])
    except Exception:
        return (0, 0)


# 픽셀 단위 컬럼 폭은 Streamlit 1.45+ 부터 지원
PX_WIDTH_OK = _st_version() >= (1, 45)

try:
    ROW_HEIGHT_OK = "row_height" in _inspect.signature(st.data_editor).parameters
except Exception:
    ROW_HEIGHT_OK = False


def _full_width_kwargs():
    """use_container_width(폐기 예정) ↔ width='stretch' 자동 선택."""
    try:
        if "width" in _inspect.signature(st.dataframe).parameters:
            return {"width": "stretch"}
    except Exception:
        pass
    return {"use_container_width": True}


FULL_W = _full_width_kwargs()


def col_width(px: int, fallback: str = "large"):
    """설치된 Streamlit 버전에 맞는 width 값 반환."""
    return px if PX_WIDTH_OK else fallback


def clean(text: str) -> str:
    """HTML 태그·엔티티 제거 후 공백 정리."""
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


PRESS_DOMAIN_MAP = {
    "hankyung.com": "한국경제", "mk.co.kr": "매일경제", "edaily.co.kr": "이데일리",
    "mt.co.kr": "머니투데이", "sedaily.com": "서울경제", "fnnews.com": "파이낸셜뉴스",
    "chosun.com": "조선일보", "biz.chosun.com": "조선비즈", "donga.com": "동아일보",
    "joongang.co.kr": "중앙일보", "joins.com": "중앙일보", "hani.co.kr": "한겨레",
    "khan.co.kr": "경향신문", "seoul.co.kr": "서울신문", "kmib.co.kr": "국민일보",
    "munhwa.com": "문화일보", "hankookilbo.com": "한국일보", "segye.com": "세계일보",
    "asiae.co.kr": "아시아경제", "ajunews.com": "아주경제", "newsis.com": "뉴시스",
    "yna.co.kr": "연합뉴스", "yonhapnews.co.kr": "연합뉴스", "news1.kr": "뉴스1",
    "heraldcorp.com": "헤럴드경제", "etnews.com": "전자신문", "dt.co.kr": "디지털타임스",
    "thebell.co.kr": "더벨", "investchosun.com": "인베스트조선", "dealsite.co.kr": "딜사이트",
    "businesspost.co.kr": "비즈니스포스트", "bizhankook.com": "비즈한국",
    "wowtv.co.kr": "한국경제TV", "moneys.co.kr": "머니S", "ceoscoredaily.com": "CEO스코어데일리",
    "housingnews.co.kr": "하우징헤럴드", "r-e.kr": "부동산일보", "kukinews.com": "쿠키뉴스",
    "newspim.com": "뉴스핌", "ebn.co.kr": "EBN", "tfmedia.co.kr": "조세금융신문",
    "g-enews.com": "글로벌이코노믹", "inews24.com": "아이뉴스24", "zdnet.co.kr": "지디넷코리아",
    "mediapen.com": "미디어펜", "econovill.com": "이코노믹리뷰", "sisajournal-e.com": "시사저널e",
    "theguru.co.kr": "더구루", "pinpointnews.co.kr": "핀포인트뉴스",
    "hankyung.co.kr": "한국경제", "seoulfn.com": "서울파이낸스",
    "financialnews.co.kr": "파이낸셜뉴스", "kbanker.co.kr": "한국금융신문",
    "fntimes.com": "한국금융신문", "insightkorea.co.kr": "인사이트코리아",
    "sisaon.co.kr": "시사오늘", "ekn.kr": "에너지경제", "m-i.kr": "매일일보",
    "smarttoday.co.kr": "스마트투데이", "newsprime.co.kr": "프라임경제",
}


def press_from_link(url: str) -> str:
    """기사 링크 도메인에서 언론사명 유추."""
    if not url:
        return PRESS_PLACEHOLDER
    try:
        host = urllib.parse.urlparse(url).netloc.lower().replace("www.", "")
        if host in PRESS_DOMAIN_MAP:
            return PRESS_DOMAIN_MAP[host]
        for domain, name in PRESS_DOMAIN_MAP.items():
            if host.endswith(domain):
                return name
    except Exception:
        pass
    return PRESS_PLACEHOLDER


# ══════════════════════════════════════════════════════════════
# 링크 해석 (리다이렉트 / 단축 URL / 네이버 뉴스)
# ══════════════════════════════════════════════════════════════
NAVER_NEWS_HOSTS = ("n.news.naver.com", "news.naver.com", "m.news.naver.com")


def resolve_final_url(url: str, timeout: int = 8) -> str:
    """리다이렉트를 따라가 최종 기사 URL 반환. HEAD 거부 시 GET으로 폴백."""
    if not url:
        return url
    try:
        r = requests.head(url, allow_redirects=True, timeout=timeout, headers=UA_HEADERS)
        if r.status_code < 400 and r.url:
            return r.url
    except Exception:
        pass
    try:
        r = requests.get(url, allow_redirects=True, timeout=timeout,
                         headers=UA_HEADERS, stream=True)
        final = r.url or url
        r.close()
        return final
    except Exception:
        return url


def extract_origin_from_naver(url: str) -> str:
    """네이버 뉴스 페이지에서 원문 링크 추출."""
    if not BS_AVAILABLE:
        return url
    try:
        r = requests.get(url, timeout=8, headers=UA_HEADERS)
        if r.status_code != 200:
            return url
        soup = BeautifulSoup(r.content, "html.parser")
        for sel, attr in [
            ('a.media_end_head_origin_link', 'href'),
            ('a[class*="origin_link"]', 'href'),
            ('meta[property="og:url"]', 'content'),
        ]:
            tag = soup.select_one(sel)
            if tag and tag.get(attr):
                cand = tag.get(attr)
                if cand and not any(h in cand for h in NAVER_NEWS_HOSTS):
                    return cand
    except Exception:
        pass
    return url


def normalize_article_url(url: str) -> str:
    """수집 링크를 실제 언론사 기사 URL로 정규화."""
    if not url:
        return url
    host = urllib.parse.urlparse(url).netloc.lower()
    if any(h in host for h in NAVER_NEWS_HOSTS):
        origin = extract_origin_from_naver(url)
        if origin and origin != url:
            return origin
        return url
    if any(k in host for k in ("bit.ly", "buly.kr", "goo.gl", "url.kr", "link.")):
        return resolve_final_url(url)
    return url


# ══════════════════════════════════════════════════════════════
# 네이버 뉴스 검색 API
# ══════════════════════════════════════════════════════════════
def fetch_naver(keyword, category, cid, csecret, hours_limit, max_pages=10, diag=None):
    rows = []
    now = dt.datetime.now(KST)
    headers = {"X-Naver-Client-Id": cid, "X-Naver-Client-Secret": csecret}
    raw_count = 0
    newest_pub = None

    for page in range(max_pages):
        start = page * 100 + 1
        if start > 900:
            break
        params = {"query": keyword, "display": 100, "start": start, "sort": "date"}
        try:
            r = requests.get("https://openapi.naver.com/v1/search/news.json",
                             headers=headers, params=params, timeout=20)
            if diag is not None:
                diag["status"] = r.status_code
            if r.status_code != 200:
                return rows, f"네이버 API 오류 {r.status_code}: {r.text[:150]}"
            items = r.json().get("items", [])
        except requests.exceptions.Timeout:
            return rows, f"네이버 API 타임아웃 (키워드: {keyword})"
        except Exception as e:
            return rows, f"네이버 요청 실패 ({keyword}): {str(e)[:100]}"

        if not items:
            break
        raw_count += len(items)

        stop = False
        for it in items:
            pub = None
            try:
                pub = dt.datetime.strptime(
                    it["pubDate"], "%a, %d %b %Y %H:%M:%S %z").astimezone(KST)
                if newest_pub is None or pub > newest_pub:
                    newest_pub = pub
            except Exception:
                pass
            if hours_limit and pub and (now - pub).total_seconds() > hours_limit * 3600:
                stop = True
                break
            link = it.get("originallink") or it.get("link", "")
            press = press_from_link(link)
            raw_title = clean(it.get("title", ""))
            rows.append({
                "카테고리": category,
                "키워드": keyword,
                # 말머리 태그·언론사 꼬리표를 떼고 순수 제목만 저장
                "제목": _clean_title(raw_title, press) or raw_title,
                "언론사": press,
                "발행시각": pub.strftime("%Y-%m-%d %H:%M") if pub else "",
                "링크": link,
                "네이버링크": it.get("link", ""),
                "요약초안": clean(it.get("description", "")),
            })
        if stop:
            break
        time.sleep(0.05)

    if diag is not None:
        diag["raw_count"] = raw_count
        diag["newest"] = newest_pub.strftime("%Y-%m-%d %H:%M") if newest_pub else "없음"
        diag["kept"] = len(rows)
    return rows, None


# ══════════════════════════════════════════════════════════════
# 중복 제거 (토큰 블로킹 + 유사도)
# ══════════════════════════════════════════════════════════════
BRACKET_RE = re.compile(r"[\[\(【〔<][^\]\)】〕>]{0,12}[\]\)】〕>]")
NONWORD_RE = re.compile(r"[^가-힣A-Za-z0-9 ]")


def normalize_title(title: str) -> str:
    t = BRACKET_RE.sub(" ", title or "")
    t = NONWORD_RE.sub(" ", t)
    return re.sub(r"\s+", " ", t).strip()


def tokens_of(title: str) -> set:
    return {w for w in normalize_title(title).split() if len(w) >= 2}


def dedup(df, title_sim_threshold=0.65, word_sim_threshold=0.6, progress_bar=None):
    """공통 토큰이 있는 기사끼리만 비교하여 전수 비교(O(n²)) 회피."""
    if df.empty:
        return df

    df = df.copy()
    df = df.drop_duplicates(subset=["링크"], keep="first")
    df = df.sort_values("발행시각", ascending=False).reset_index(drop=True)

    kept_rows, kept_norm, kept_tokens = [], [], []
    token_index = {}
    total = len(df)

    for i, (_, row) in enumerate(df.iterrows()):
        title = str(row["제목"])
        norm = normalize_title(title)
        toks = tokens_of(title)

        candidates = set()
        for t in toks:
            candidates.update(token_index.get(t, ()))

        is_dup = False
        for ci in candidates:
            kt = kept_tokens[ci]
            union = len(toks | kt)
            if union == 0:
                continue
            word_sim = len(toks & kt) / union
            if word_sim >= word_sim_threshold:
                is_dup = True
                break
            if word_sim >= 0.3:
                if SequenceMatcher(None, norm, kept_norm[ci]).ratio() >= title_sim_threshold:
                    is_dup = True
                    break

        if not is_dup:
            idx = len(kept_rows)
            kept_rows.append(row)
            kept_norm.append(norm)
            kept_tokens.append(toks)
            for t in toks:
                token_index.setdefault(t, []).append(idx)

        if progress_bar is not None and total > 0 and (i % 20 == 0 or i == total - 1):
            progress_bar.progress(min((i + 1) / total, 1.0),
                                  text=f"중복 제거 중... ({i + 1}/{total})")

    return pd.DataFrame(kept_rows).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════
# 관련성 필터 (상업용 부동산 / 기관투자 / 부동산금융 / 정책)
# ══════════════════════════════════════════════════════════════
# 제목에 있으면 무조건 관련 기사로 취급하는 강신호
STRONG_POS = (
    "오피스", "물류센터", "데이터센터", "지식산업센터", "리츠", "자산운용",
    "매각주관", "사옥", "상업용", "공실률", "캡레이트", "부동산펀드",
    "기관투자자", "연기금", "임대차", "프라임", "빌딩", "복합개발",
)

POS_TERMS = [
    ("오피스", 4), ("물류센터", 4), ("데이터센터", 4), ("지식산업센터", 4),
    ("리츠", 4), ("자산운용", 4), ("매각주관", 4), ("사옥", 4), ("상업용", 4),
    ("공실", 4), ("캡레이트", 4), ("부동산펀드", 4), ("기관투자자", 4),
    ("연기금", 4), ("임대차", 4), ("프라임", 3), ("빌딩", 3), ("복합개발", 3),
    ("리테일", 3), ("상권", 3), ("호텔", 2), ("임대료", 3), ("수익률", 2),
    ("우선협상", 3), ("클로징", 3), ("실사", 2), ("주관사", 3), ("신탁", 3),
    ("PF", 3), ("프로젝트파이낸싱", 3), ("브릿지론", 3), ("메자닌", 3),
    ("블라인드펀드", 3), ("운용사", 3), ("투자자", 2), ("리모델링", 2),
    ("착공", 2), ("준공", 2), ("인허가", 2), ("부지", 2), ("본사 이전", 3),
    ("임차", 3), ("입주", 2), ("매각", 2), ("매입", 2), ("인수", 1),
    ("거래", 1), ("펀드", 2), ("국토부", 2), ("국토교통부", 2), ("세제", 2),
    ("규제 완화", 2), ("부동산 정책", 3), ("금융당국", 2), ("보험사", 1),
    ("증권사", 1), ("개발사업", 3), ("시행사", 3), ("디벨로퍼", 3),
]

NEG_TERMS = [
    # 주거용 · 개인 재테크
    ("분양", 5), ("청약", 5), ("입주자모집", 5), ("전세사기", 5), ("갭투자", 5),
    ("빌라", 4), ("다세대", 4), ("아파트값", 5), ("집값", 4), ("전셋값", 4),
    ("분양가", 4), ("모델하우스", 5), ("견본주택", 5), ("신혼부부", 4),
    ("임대주택", 3), ("공공주택", 3), ("아파트", 3), ("주택담보", 3),
    ("재건축", 2), ("재개발", 2), ("월세", 2), ("이사", 2),
    # 무관 분야
    ("연예", 5), ("배우", 5), ("가수", 5), ("아이돌", 5), ("드라마", 4),
    ("예능", 5), ("프로야구", 5), ("축구", 4), ("골프", 3), ("날씨", 5),
    ("코인", 4), ("비트코인", 5), ("특징주", 5), ("상한가", 5), ("급등주", 5),
    ("종목추천", 5), ("부고", 5), ("화보", 5), ("음주운전", 5), ("마약", 5),
    ("성폭행", 5), ("이혼", 4), ("복권", 5), ("로또", 5), ("맛집", 5),
    ("건강", 3), ("다이어트", 5), ("채용공고", 4), ("장학금", 4),
]


def relevance_score(title: str, desc: str = "") -> int:
    """제목 가중 2배 + 발췌문으로 상업용 부동산 관련도 점수 산출."""
    t = str(title or "")
    text = f"{t} {t} {str(desc or '')}"
    pos = sum(w for term, w in POS_TERMS if term in text)
    neg = sum(w for term, w in NEG_TERMS if term in text)
    score = pos - neg
    if any(term in t for term in STRONG_POS):
        score += 4          # 핵심 자산군이 제목에 있으면 함부로 버리지 않는다
    return int(score)


def ensure_filter_cols(df):
    """이전 버전 세션 데이터에도 관련성 컬럼을 보강 (KeyError 방지)."""
    if df is None or len(df) == 0:
        return df
    if "관련도" not in df.columns:
        descs = df["요약초안"] if "요약초안" in df.columns else [""] * len(df)
        df["관련도"] = [relevance_score(t, d) for t, d in zip(df["제목"], descs)]
    if "제외" not in df.columns:
        df["제외"] = False
    if "제외사유" not in df.columns:
        df["제외사유"] = ""
    df["제외"] = df["제외"].fillna(False).astype(bool)
    return df


FILTER_LEVELS = {
    "약하게 (명백한 것만 제외)": -6,
    "보통 (권장)": -1,
    "강하게 (애매하면 제외)": 3,
}

# 이 점수 미만이면 AI 재판정 대상 (명백히 관련된 기사는 호출 낭비 방지)
AI_CHECK_BELOW = 8

AI_RELEVANCE_PROMPT = """아래는 뉴스 제목 목록이다. 이 중에서 '상업용 부동산' 업계
종사자에게 의미 있는 기사만 골라라.

관련 있음: 오피스·물류센터·데이터센터·리테일·호텔 등 상업용 부동산의 매매·개발·
임대차·시장동향 / 리츠·부동산펀드·자산운용사·기관투자자의 부동산 투자 /
부동산 PF·신탁·대출 등 부동산금융 / 부동산 관련 정부 정책·규제·세제

관련 없음: 아파트 분양·청약·전세 등 주거용 / 개인 재테크 / 부동산과 무관한
기업 실적·인사 / 연예·스포츠·사건사고 / 주식 종목 기사

목록:
{items}

관련 있는 항목의 번호만 쉼표로 구분해 출력하라. 다른 말은 쓰지 마라.
예시 출력: 1,3,4,9"""


def ai_relevance_check(titles, gemini_key, picker, batch=40):
    """
    제목을 묶어 Gemini에 관련성 판정 요청.
    반환: (관련 있다고 판정된 위치 집합, 에러목록)
    판정 실패한 배치는 전부 '관련'으로 간주해 기사를 잃지 않는다.
    """
    keep, errors = set(), []
    for start in range(0, len(titles), batch):
        chunk = titles[start:start + batch]
        items = "\n".join(f"{i + 1}. {t[:80]}" for i, t in enumerate(chunk))
        prompt = AI_RELEVANCE_PROMPT.format(items=items)

        raw, err = "", None
        for _ in range(max(len(picker.candidates), 1)):
            model = picker.current()
            if not model:
                break
            raw, err, dead = _call_gemini(prompt, gemini_key, model)
            if dead:
                picker.mark_dead(model, err)
                continue
            break

        if not raw:
            errors.append(err or "응답 없음")
            keep.update(range(start, start + len(chunk)))   # 실패 시 전부 유지
            continue

        nums = {int(n) for n in re.findall(r"\d{1,3}", raw)}
        picked = {start + n - 1 for n in nums if 1 <= n <= len(chunk)}
        if not picked:      # 전부 무관이라는 응답은 오작동 가능성 → 유지
            errors.append("판정 결과 없음 (전체 유지)")
            picked = set(range(start, start + len(chunk)))
        keep.update(picked)

    return keep, errors


# ══════════════════════════════════════════════════════════════
# 페이지 수집
# ══════════════════════════════════════════════════════════════
def fetch_html(url: str, timeout: int = 10) -> str:
    """기사 HTML 1회만 받아 bs4·trafilatura·언론사·제목 추출이 함께 사용."""
    try:
        r = requests.get(url, timeout=timeout, headers=UA_HEADERS)
        if r.status_code != 200:
            return ""
        r.encoding = r.apparent_encoding or r.encoding
        return r.text
    except Exception:
        return ""


# ── 언론사명 추출 ─────────────────────────────────────────────
def _clean_press_name(name: str) -> str:
    if not name:
        return ""
    name = html.unescape(str(name)).strip().strip('"\'|-·<>[]')
    name = re.sub(r"\s+", " ", name)
    name = re.sub(r"^@", "", name)
    name = re.sub(r"\s*(홈페이지|공식|모바일|온라인|PC)$", "", name).strip()
    if not name or len(name) > 20:
        return ""
    low = name.lower()
    if any(bad in low for bad in ("네이버", "naver", "다음", "daum", "google", "포털")):
        return ""
    if re.fullmatch(r"[a-z0-9.\-_/]+", low) and "." in low:
        return ""
    return name


PRESS_META_SELECTORS = [
    ('meta[property="og:site_name"]', "content"),
    ('meta[name="og:site_name"]', "content"),
    ('meta[property="dable:author"]', "content"),
    ('meta[name="article:media_name"]', "content"),
    ('meta[name="twitter:site"]', "content"),
    ('meta[name="publisher"]', "content"),
    ('meta[property="article:publisher_name"]', "content"),
]


def press_from_html(soup) -> str:
    """og:site_name → JSON-LD publisher → <title> 꼬리표 순으로 언론사명 추출."""
    for sel, attr in PRESS_META_SELECTORS:
        tag = soup.select_one(sel)
        if tag:
            v = _clean_press_name(tag.get(attr, ""))
            if v:
                return v

    for script in soup.select('script[type="application/ld+json"]'):
        try:
            txt = script.string or script.get_text() or ""
        except Exception:
            continue
        m = re.search(r'"publisher"\s*:\s*\{[^{}]*?"name"\s*:\s*"([^"]{1,40})"', txt, re.S)
        if m:
            v = _clean_press_name(m.group(1))
            if v:
                return v

    try:
        title = soup.title.string if soup.title else ""
    except Exception:
        title = ""
    if title:
        for sep in (" - ", " | ", " < ", " :: ", " > ", " – "):
            if sep in title:
                v = _clean_press_name(title.rsplit(sep, 1)[-1])
                if v:
                    return v
    return ""


def press_fallback_from_url(url: str) -> str:
    """매핑·HTML 모두 실패 시 도메인이라도 노출."""
    try:
        host = urllib.parse.urlparse(url).netloc.lower().replace("www.", "")
        return host or PRESS_PLACEHOLDER
    except Exception:
        return PRESS_PLACEHOLDER


def resolve_press(final_url: str, raw_html: str = "") -> str:
    """도메인 매핑(정식 명칭) 우선 → HTML 메타 → 도메인."""
    mapped = press_from_link(final_url)
    if mapped != PRESS_PLACEHOLDER:
        return mapped
    if raw_html and BS_AVAILABLE:
        try:
            v = press_from_html(BeautifulSoup(raw_html, "html.parser"))
            if v:
                return v
        except Exception:
            pass
    return press_fallback_from_url(final_url)


# ── 원문 제목 추출 ────────────────────────────────────────────
TRUNC_RE = re.compile(r"(\.\.\.+|…+|‥+)\s*$")
TITLE_TAIL_SEPS = (" - ", " | ", " < ", " :: ", " > ", " – ", " — ")

# 제목 앞뒤에 붙는 말머리 태그: [유교신문], (종합), 【단독】 …
LEAD_TAG_RE = re.compile(r"^\s*[\[\(<【〔]\s*([^\]\)>】〕]{1,24})\s*[\]\)>】〕]\s*")
TAIL_TAG_RE = re.compile(r"\s*[\[\(<【〔]\s*([^\]\)>】〕]{1,24})\s*[\]\)>】〕]\s*$")

PRESS_NAME_SET = set(PRESS_DOMAIN_MAP.values())
PRESS_LIKE_RE = re.compile(
    r"(신문|일보|경제|뉴스|타임스|저널|투데이|미디어|방송|데일리|헤럴드|포스트|"
    r"통신|리포트|매거진|TV|News|Times|Post|Daily)$", re.I)


def _is_press_like(s: str) -> bool:
    s = (s or "").strip()
    if not s or len(s) > 15:
        return False
    return s in PRESS_NAME_SET or bool(PRESS_LIKE_RE.search(s))


def strip_decor_tags(title: str) -> str:
    """제목 앞뒤의 말머리 태그를 모두 제거 ([유교신문], [단독], (종합), <상> 등)."""
    t = (title or "").strip()
    for _ in range(5):
        before = t
        m = LEAD_TAG_RE.match(t)
        if m:
            cand = t[m.end():].strip()
            if len(cand) >= 10:
                t = cand
        m = TAIL_TAG_RE.search(t)
        if m:
            cand = t[:m.start()].strip()
            if len(cand) >= 10:
                t = cand
        if t == before:
            break
    return t


def strip_press_suffix(title: str, press: str = "") -> str:
    """'제목 | 연합뉴스' 처럼 꼬리에 붙은 언론사명 제거."""
    t = (title or "").strip()
    for _ in range(2):
        before = t
        for sep in TITLE_TAIL_SEPS:
            if sep in t:
                head, tail = t.rsplit(sep, 1)
                head, tail = head.strip(), tail.strip()
                if len(head) >= 10 and (
                        (press and tail == press.strip()) or _is_press_like(tail)):
                    t = head
                    break
        if t == before:
            break
    return t


def _clean_title(t: str, press: str = "") -> str:
    """공백 정리 → 언론사 꼬리표 제거 → 말머리 태그 제거."""
    if not t:
        return ""
    t = re.sub(r"\s+", " ", html.unescape(str(t))).strip().strip("\"'“”‘’")
    t = strip_press_suffix(t, press)
    t = strip_decor_tags(t)
    t = strip_press_suffix(t, press)
    return t[:200] if len(t) >= 8 else ""


def title_from_html(soup, press: str = "") -> str:
    """og:title → twitter:title → h1 → <title> 순으로 기사 원문 제목 추출."""
    for sel, attr in [('meta[property="og:title"]', "content"),
                      ('meta[name="twitter:title"]', "content"),
                      ('meta[name="title"]', "content")]:
        tag = soup.select_one(sel)
        if tag:
            v = _clean_title(tag.get(attr, ""), press)
            if v and not TRUNC_RE.search(v):
                return v

    for sel in ("h1.headline", "h1#title", "h1", ".article-head-title", "#articleTitle"):
        tag = soup.select_one(sel)
        if tag:
            v = _clean_title(tag.get_text(), press)
            if v and not TRUNC_RE.search(v):
                return v

    try:
        raw = soup.title.string if soup.title else ""
    except Exception:
        raw = ""
    v = _clean_title(raw, press)
    return "" if TRUNC_RE.search(v) else v


def better_title(current: str, crawled: str) -> str:
    """네이버 제목이 잘려 있으면 원문 제목으로 교체."""
    cur = (current or "").strip()
    new = (crawled or "").strip()
    if not new or new == cur:
        return cur
    if not cur:
        return new

    core = TRUNC_RE.sub("", cur).strip()
    if len(core) < 6:
        return cur
    same_article = new.startswith(core[:12]) or core[:12] in new

    if TRUNC_RE.search(cur) and len(new) > len(core):
        return new
    if same_article and len(new) > len(cur) + 4:
        return new
    return cur


# ── 본문 파싱 ────────────────────────────────────────────────
CAPTION_RE = re.compile(
    r"^\s*(?:[▲◀▶★●○□■▼△▽◇◆※☞]"
    r"|\[?\s*(?:사진|영상|자료|출처|제공|그래픽|표|이미지)\s*[=:\]]"
    r"|\(\s*(?:사진|영상|자료|출처|제공)\s*[=:]?)"
)
REPORTER_RE = re.compile(
    r"(기자\s*[=:]|무단\s*전재|재배포\s*금지|저작권자|ⓒ|Copyright|@[\w.]+\.(?:co\.kr|com|kr))")

ARTICLE_SELECTORS = [
    "#dic_area", "#newsct_article", "#articleBodyContents", "#articeBody",  # 네이버
    "#article-view-content-div", ".article-body", ".article_body", ".articleBody",
    ".news-body", ".article_content", ".article-content", ".articleText",
    ".news-content", ".entry-content", "#articleBody", "#news_body_area",
    "#CmAdContent", "article", "#content", "main",
]

DROP_SELECTORS = [
    "script", "style", "nav", "footer", "header", "aside", "iframe",
    ".nav", ".menu", ".ad", ".advertisement", ".comment", ".related",
    ".sidebar", ".social", "figure", "figcaption", ".caption",
    ".photo-caption", ".reporter", ".copyright", ".byline",
]


def _clean_paragraphs(lines):
    out = []
    for line in lines:
        line = line.strip()
        if len(line) < 15:
            continue
        if CAPTION_RE.match(line):
            continue
        if REPORTER_RE.search(line) and len(line) < 80:
            continue
        out.append(line)
    return out


def parse_body_with_bs4(raw_html: str):
    """반환: (본문, 언론사, 원문제목)"""
    if not BS_AVAILABLE or not raw_html:
        return "", "", ""
    try:
        soup = BeautifulSoup(raw_html, "html.parser")
        # decompose 전에 먼저 추출 (h1/header가 제거 대상에 포함되므로)
        press = press_from_html(soup)
        page_title = title_from_html(soup, press)

        for selector in DROP_SELECTORS:
            for el in soup.select(selector):
                el.decompose()

        body = None
        for selector in ARTICLE_SELECTORS:
            body = soup.select_one(selector)
            if body:
                break

        target = body if body is not None else soup
        paragraphs = target.find_all("p")
        if paragraphs and len(" ".join(p.get_text() for p in paragraphs)) > 200:
            lines = [p.get_text() for p in paragraphs]
        else:
            lines = target.get_text("\n").split("\n")

        text = re.sub(r"\s+", " ", " ".join(_clean_paragraphs(lines))).strip()
        return (text[:2500] if len(text) >= 150 else ""), press, page_title
    except Exception:
        return "", "", ""


def parse_body_with_trafilatura(raw_html: str) -> str:
    if not trafilatura or not raw_html:
        return ""
    try:
        text = trafilatura.extract(raw_html, include_comments=False, include_tables=False)
        if not text:
            return ""
        text = re.sub(r"\s+", " ", " ".join(_clean_paragraphs(text.split("\n")))).strip()
        return text[:2500] if len(text) >= 150 else ""
    except Exception:
        return ""


def extract_text_with_newspaper(url: str, timeout: int = 8) -> str:
    if not NEWSPAPER_AVAILABLE:
        return ""
    try:
        cfg = NewspaperConfig()
        cfg.browser_user_agent = UA_HEADERS["User-Agent"]
        cfg.request_timeout = timeout
        cfg.fetch_images = False
        cfg.memoize_articles = False
        article = Article(url, language="ko", config=cfg)
        article.download()
        article.parse()
        text = " ".join(_clean_paragraphs((article.text or "").split("\n")))
        text = re.sub(r"\s+", " ", text).strip()
        return text[:2500] if len(text) >= 150 else ""
    except Exception:
        return ""


def fetch_article(url: str):
    """
    링크 정규화 → HTML 1회 수집 → 3단계 폴백 본문 추출 + 언론사·원문제목 판별.
    반환: (본문, 사용된추출기, 최종URL, 언론사, 원문제목)
    """
    final_url = normalize_article_url(url)
    text, press, page_title, extractor = "", "", "", "실패"

    raw = fetch_html(final_url)
    if raw:
        text, press, page_title = parse_body_with_bs4(raw)
        if text:
            extractor = "bs4"
        else:
            text = parse_body_with_trafilatura(raw)
            if text:
                extractor = "trafilatura"

    if not text:
        text = extract_text_with_newspaper(final_url)
        if text:
            extractor = "newspaper"

    if not text and final_url != url:
        raw2 = fetch_html(url)
        if raw2:
            text, p2, t2 = parse_body_with_bs4(raw2)
            press = press or p2
            page_title = page_title or t2
            if text:
                extractor, final_url = "bs4(원본)", url

    return text, extractor, final_url, resolve_press(final_url, raw), page_title


def first_sentence(text: str, max_chars: int = 150) -> str:
    if not text:
        return ""
    m = re.search(r"[^.!?\n]*[.!?]", text)
    return (m.group(0) if m else text[:max_chars].rstrip() + ".")[:max_chars].strip()


def naver_fallback_summary(desc: str, max_chars: int = 90) -> str:
    """
    본문 크롤링이 막힌 매체(더벨·인베스트조선 등) 대비.
    네이버 검색 API가 함께 주는 요약문을 정리해 사용한다.
    """
    t = re.sub(r"\s+", " ", str(desc or "")).strip()
    if not t:
        return ""
    t = TRUNC_RE.sub("", t).strip()      # 끝의 '...' 제거
    if len(t) < 12:
        return ""
    return first_sentence(t, max_chars)


# ══════════════════════════════════════════════════════════════
# Gemini (REST 직접 호출)
# ══════════════════════════════════════════════════════════════
GEMINI_PROMPT = """아래 기사를 부동산 업계용 '기사 상단 요약'으로 바꾸고 분류하라.

출력 형식(이 두 줄 외에는 아무것도 쓰지 말 것):
분류: <아래 6개 중 하나>
요약: <1줄. 한 줄에 안 담기면 줄바꿈 후 2줄까지>

분류 선택지와 판단 기준:
- 매입매각: 자산·지분·건물의 매각·매입·인수, 매각주관사 선정, 우선협상자, 딜 클로징
- 개발계획: 신축·복합개발·착공·준공·인허가·부지확보·설계·기공
- 이전동향: 사옥·본사 이전, 신규 임차, 입주, 리모델링, 사무실 이동
- 시장동향: 공실률·임대료·수익률·거래량·가격지수 등 시장 통계와 전망
- 정책: 정부·국토부·금융당국의 규제, 세제, 법·제도 개정
- 업계동향: 운용사·증권·보험·건설사의 실적·인사·조직, 펀드·리츠 설정, 그 외

요약 작성 조건:
각 줄 35~65자. 모든 줄을 명사로 끝낼 것(매각·매입·추진·확정·완료·착수·예정·
검토·전망·체결 등). '했다/한다/이다/밝혔다' 같은 서술형 어미 금지.
첫 줄은 주체+대상+규모+행위. 둘째 줄은 기사에 실제로 있는 내용만.
번호·불릿·따옴표 없이.

좋은 예:
분류: 매입매각
요약: 현대건설, 여의도 사옥 4500억 규모 매각 완료
매수자는 이지스자산운용, 평당 3200만원으로 3분기 서울 오피스 최대 거래

--- 기사 ---
{text}
--- 출력 ---"""

STRICT_SUFFIX = "\n\n(직전 답변에 서술형 어미가 있었다. 모든 줄을 명사로 끝내라.)"

GEMINI_SHORT_PROMPT = """아래는 기사 전문이 아니라 검색 결과에 노출된 짧은 발췌문이다.
제목과 발췌문에 실제로 있는 사실만으로 1줄 요약과 분류를 작성하라.
없는 내용은 절대 지어내지 마라.

출력 형식(이 두 줄 외에는 아무것도 쓰지 말 것):
분류: <매입매각|개발계획|이전동향|시장동향|정책|업계동향 중 하나>
요약: <1줄. 30~65자. 반드시 명사로 종결. '했다/한다/이다/전해진다' 금지>

분류 기준: 매입매각=매각·매입·인수·주관사·우선협상 / 개발계획=신축·착공·준공·인허가·
부지 / 이전동향=사옥·본사 이전·임차·입주 / 시장동향=공실률·임대료·수익률·거래량·전망 /
정책=정부·국토부·규제·세제·법개정 / 업계동향=실적·인사·펀드·리츠 설정·그 외

제목: {title}
발췌문: {text}
--- 출력 ---"""

CATEGORY_LINE_RE = re.compile(r"^\s*(?:분류|카테고리|category)\s*[:：]\s*(.+)$", re.I)
SUMMARY_PREFIX_RE = re.compile(r"^\s*(?:요약|summary)\s*[:：]\s*", re.I)

VERB_END_RE = re.compile(
    r"(했다|한다|이다|된다|였다|겠다|봤다|섰다|왔다|났다|한다고|합니다|입니다|습니다|"
    r"보인다|늘었다|줄었다|것이다|밝혔다|전했다|나타났다)\s*[.]?$")

BULLET_RE = re.compile(r"^\s*(?:[-•*▷▶·ㆍ○●]|\d+[.)])\s*")

# 프롬프트 규칙문이나 모델의 사고 과정이 새어 나온 줄을 걸러내기 위한 패턴
PROMPT_LEAK_RE = re.compile(
    r"(줄차|명사형|명사로|동사형|서술형|기사에 실제로|실제로 있는 것만|작성 규칙|"
    r"헤드라인만|요약문만|출력 금지|추측 금지|줄바꿈으로만|거래상대방, 단가|"
    r"예시\(|좋은 예|경제지 기자|최대 2줄|어미|규칙 \d|사용자|요약해야|"
    r"해야 한다|끝낼 것|^조건|^-{2,}|기사 ---|출력 형식|분류 선택지|판단 기준|"
    r"작성 조건|아래 6개|선택지와|발췌문|지어내지|분류 기준)")

# 프롬프트가 35~65자를 요구하므로, 이보다 크게 짧은 줄은 문장 조각으로 간주
MIN_SUMMARY_LEN = 18


def _valid_summary_line(ln: str) -> bool:
    """문장 조각·프롬프트 유출 줄 배제."""
    if len(ln) < MIN_SUMMARY_LEN:
        return False
    if PROMPT_LEAK_RE.search(ln):
        return False
    if ln[0] in ",·:;)]}…”’-+/=":
        return False
    if ln.rstrip().endswith((",", "·", "및", "와", "과", "의")):
        return False
    if not re.search(r"[가-힣]", ln):
        return False
    return True


def _postprocess_summary(raw: str):
    """
    '분류:' 줄과 요약 줄을 분리하고 불릿·프롬프트 유출·문장 조각을 제거.
    반환: (카테고리, 요약, 동사형발견여부)
    """
    category, lines = "", []
    for ln in (raw or "").split("\n"):
        ln = ln.strip().strip("*#").strip()          # 마크다운 강조 제거
        if not ln:
            continue

        m = CATEGORY_LINE_RE.match(ln)
        if m and not category:
            cand = re.sub(r"[^가-힣]", "", m.group(1))
            for c in MAIL_CATEGORIES:
                if c in cand:
                    category = c
                    break
            continue

        ln = SUMMARY_PREFIX_RE.sub("", ln)
        ln = BULLET_RE.sub("", ln).strip(" \"'`“”‘’")
        ln = re.sub(r"\s+", " ", ln)
        if _valid_summary_line(ln):
            lines.append(ln[:90])
        if len(lines) == 2:
            break

    if not lines:
        return category, "", False
    return category, "\n".join(lines), any(VERB_END_RE.search(l) for l in lines)


BAD_MODEL_TOKENS = ("embedding", "aqa", "vision", "imagen", "tts", "live",
                    "gemma", "image", "veo", "learnlm", "thinking")


def model_score(n: str) -> float:
    """버전 숫자를 직접 파싱해 신모델이 나와도 자동으로 상위 랭크."""
    low = n.lower()
    s = 0.0
    if "flash" in low:
        s += 20
    if "pro" in low:
        s += 8
    if "lite" in low:
        s -= 10          # 404 다발 → 후순위
    if "latest" in low:
        s += 5
    if "exp" in low or "preview" in low or re.search(r"-\d{3,4}$", low):
        s -= 8
    m = re.search(r"(\d+)\.(\d+)", low)
    if m:
        s += int(m.group(1)) * 3 + int(m.group(2)) * 0.3
    else:
        s += 4
    return s


def list_gemini_models(gemini_key: str):
    """generateContent 가능한 모델을 선호도 순으로 반환. 반환: (모델리스트, 에러)"""
    if not gemini_key:
        return [], "GEMINI_API_KEY 없음"
    try:
        r = requests.get(
            "https://generativelanguage.googleapis.com/v1beta/models",
            headers={"x-goog-api-key": gemini_key}, timeout=15)
    except Exception as e:
        return [], f"모델목록 조회 오류: {str(e)[:120]}"

    if r.status_code != 200:
        return [], f"모델목록 조회 실패 HTTP {r.status_code}: {r.text[:150]}"

    usable = []
    for m in r.json().get("models", []):
        name = m.get("name", "").replace("models/", "")
        if "generateContent" not in m.get("supportedGenerationMethods", []):
            continue
        if any(bad in name.lower() for bad in BAD_MODEL_TOKENS):
            continue
        usable.append(name)
    if not usable:
        return [], "generateContent 지원 모델 없음"

    usable.sort(key=model_score, reverse=True)
    return usable, None


class ModelPicker:
    """
    listModels에는 뜨지만 generateContent는 404를 주는 모델이 존재한다.
    실제 호출이 실패한 모델을 즉시 제외하고 다음 후보로 넘어간다. 스레드 안전.
    """

    def __init__(self, candidates):
        self.candidates = [c for c in candidates if c]
        self.dead = {}
        self._lock = threading.Lock()

    def current(self):
        with self._lock:
            for m in self.candidates:
                if m not in self.dead:
                    return m
        return ""

    def mark_dead(self, model, reason=""):
        with self._lock:
            if model and model not in self.dead:
                self.dead[model] = reason

    def alive(self):
        with self._lock:
            return [m for m in self.candidates if m not in self.dead]


def resolve_gemini_candidates(gemini_key: str):
    """성공 결과만 세션 캐시. 반환: (후보리스트, 에러)"""
    cached = st.session_state.get("_gemini_model_list")
    if cached:
        return cached, None
    models, err = list_gemini_models(gemini_key)
    if not models:
        return [], err
    st.session_state["_gemini_model_list"] = models
    st.session_state.setdefault("_gemini_model", models[0])
    return models, None


DEAD_MODEL_CODES = (400, 403, 404)
MAX_OUTPUT_TOKENS = 800

# 429 응답에서 원인·대기시간을 뽑아내기 위한 패턴
RETRY_DELAY_RE = re.compile(r'"retryDelay"\s*:\s*"(\d+(?:\.\d+)?)s"')
QUOTA_ID_RE = re.compile(r'"quotaId"\s*:\s*"([^"]+)"')
QUOTA_METRIC_RE = re.compile(r'"quotaMetric"\s*:\s*"([^"]+)"')


def _quota_reason(body: str):
    """429 본문에서 (사람이 읽을 원인, 일일한도여부) 추출."""
    q = QUOTA_ID_RE.search(body) or QUOTA_METRIC_RE.search(body)
    qid = q.group(1) if q else ""
    per_day = bool(re.search(r"PerDay|per_day", qid, re.I))
    if qid:
        kind = "일일 한도(RPD)" if per_day else "분당 한도(RPM/TPM)"
        return f"{kind} · {qid.split('/')[-1]}", per_day
    return "원인 미상", False


def _call_gemini(prompt: str, gemini_key: str, model_name: str):
    """
    단일 호출 + 재시도. 반환: (원문텍스트, 에러, 모델폐기여부)

    2.5 이상 모델은 기본이 '사고(thinking) 모드'라 출력 토큰을 내부 추론에
    소진한다. → thinkingBudget=0 으로 끄고, thought 파트는 결과에서 제외하며,
    MAX_TOKENS 종료는 실패로 처리한다.
    """
    headers = {"Content-Type": "application/json", "x-goog-api-key": gemini_key}
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent"

    def build(thinking_off: bool):
        gc = {"maxOutputTokens": MAX_OUTPUT_TOKENS, "temperature": 0.3}
        if thinking_off:
            gc["thinkingConfig"] = {"thinkingBudget": 0}
        return {"contents": [{"parts": [{"text": prompt}]}], "generationConfig": gc}

    thinking_off = True
    last_err = ""

    for attempt in range(4):
        try:
            r = requests.post(url, headers=headers, json=build(thinking_off), timeout=30)
        except Exception as e:
            last_err = f"{model_name}: 요청 오류 {str(e)[:100]}"
            time.sleep(1.5)
            continue

        if r.status_code == 200:
            data = r.json()
            candidates = data.get("candidates", [])
            if not candidates:
                fb = data.get("promptFeedback", {})
                return "", f"{model_name}: 후보 없음 (blockReason={fb.get('blockReason', '?')})", False

            cand = candidates[0]
            finish = cand.get("finishReason", "")
            parts = cand.get("content", {}).get("parts", [])
            out = "".join(p.get("text", "") for p in parts if not p.get("thought")).strip()

            if finish == "MAX_TOKENS":
                if thinking_off:
                    return "", f"{model_name}: 출력 토큰 초과 (응답 잘림)", False
                thinking_off = True
                continue
            if out:
                return out, None, False
            return "", f"{model_name}: 빈 응답 (finishReason={finish or '?'})", False

        if r.status_code == 429:
            body = re.sub(r"\s+", " ", r.text)
            reason, per_day = _quota_reason(body)
            # 일일 한도는 기다려도 안 풀린다. 단 모델별로 따로 집계되므로
            # 이 모델만 폐기 처리하고 다음 후보 모델로 넘긴다.
            if per_day:
                return "", (f"{model_name}: 오늘 무료 일일 한도 소진 ({reason}) "
                            f"→ 다른 모델로 전환"), True
            m = RETRY_DELAY_RE.search(body)
            delay = min(float(m.group(1)) if m else 8 * (attempt + 1), 60)
            last_err = (f"{model_name}: 429 {reason} · "
                        f"{delay:.0f}초 대기 후 재시도 {attempt + 1}/4")
            time.sleep(delay)
            continue

        if r.status_code in (500, 503):
            last_err = f"{model_name} HTTP {r.status_code} (일시 오류, 재시도)"
            time.sleep(2.0)
            continue

        msg = re.sub(r"\s+", " ", r.text)[:180]

        # thinkingConfig를 지원하지 않는 모델 → 옵션 빼고 재시도 (폐기하지 않음)
        if r.status_code == 400 and thinking_off and re.search(r"thinking", msg, re.I):
            thinking_off = False
            continue

        return "", f"{model_name} HTTP {r.status_code}: {msg}", r.status_code in DEAD_MODEL_CODES

    return "", last_err or f"{model_name}: 알 수 없는 오류", False


def generate_summary_with_gemini(article_text: str, gemini_key: str,
                                 picker: "ModelPicker", title: str = "",
                                 short: bool = False):
    """
    명사형 1~2줄 요약 + 메일 카테고리 분류. 사용 불가 모델(404 등)은 자동 폴백.
    동사형 어미가 섞이면 1회 재요청. 반환: (카테고리, 요약, 에러)
    """
    if not gemini_key:
        return "", "", "GEMINI_API_KEY 없음"
    if not article_text:
        return "", "", "본문 없음"
    if picker is None or not picker.current():
        return "", "", "사용 가능한 Gemini 모델 없음"

    if short:   # 본문 대신 네이버 검색 발췌문만 있는 경우
        prompt = GEMINI_SHORT_PROMPT.format(title=str(title)[:120],
                                            text=article_text[:1200])
    else:
        prompt = GEMINI_PROMPT.format(text=article_text[:2500])
    errors = []

    # 후보 전체를 순회 (일일 한도는 모델별이라 다른 모델은 살아 있을 수 있다)
    for _ in range(max(len(picker.candidates), 1)):
        model_name = picker.current()
        if not model_name:
            break

        raw, err, dead = _call_gemini(prompt, gemini_key, model_name)
        if dead:
            picker.mark_dead(model_name, err)
            errors.append(err)
            continue
        if not raw:
            return "", "", err

        category, summary, has_verb = _postprocess_summary(raw)
        if has_verb:
            raw2, _e2, _d2 = _call_gemini(prompt + STRICT_SUFFIX, gemini_key, model_name)
            if raw2:
                cat2, summary2, has_verb2 = _postprocess_summary(raw2)
                if summary2 and not has_verb2:
                    return (cat2 or category), summary2, None
        if summary:
            return category, summary, None
        return category, "", f"{model_name}: 후처리 후 빈 결과"

    if errors and all("일일 한도" in e for e in errors):
        return "", "", f"무료 일일 한도 소진 (모델 {len(errors)}개 전부)"
    return "", "", "모든 모델 사용 불가 · " + " / ".join(errors[:2])


def summarize_one(row_idx, url, title, desc, gemini_key, picker, use_ai):
    """
    워커: 본문 추출 + 언론사·원문제목 판별 + 요약·분류 생성.
    본문이 막힌 매체(더벨 등)는 네이버 검색 발췌문을 Gemini에 넘겨 요약한다.
    반환: (idx, 카테고리, 요약, 언론사, 원문제목, 로그)
    """
    text, extractor, final_url, press, page_title = fetch_article(url)

    # ── 본문 확보 실패 → 네이버 검색 발췌문으로 대체 ──
    if not text:
        excerpt = TRUNC_RE.sub("", re.sub(r"\s+", " ", str(desc or "")).strip()).strip()
        if use_ai and len(excerpt) >= 30:
            category, summary, err = generate_summary_with_gemini(
                excerpt, gemini_key, picker, title=title, short=True)
            if summary:
                return (row_idx, category, summary, press, page_title,
                        f"본문 차단({final_url[:45]}) → 네이버 발췌문으로 AI 요약")
            return (row_idx, category, naver_fallback_summary(excerpt), press,
                    page_title, f"본문 차단 + AI 실패 → 발췌문 원문 사용 ({err})")
        return (row_idx, "", naver_fallback_summary(excerpt), press, page_title,
                f"본문 추출 실패 → {final_url[:70]}")

    if not use_ai:
        return row_idx, "", first_sentence(text), press, page_title, None

    category, summary, err = generate_summary_with_gemini(
        text, gemini_key, picker, title=title)
    if summary:
        return row_idx, category, summary, press, page_title, None
    return row_idx, category, first_sentence(text), press, page_title, \
        f"[Gemini실패/{extractor}] {err}"


# ══════════════════════════════════════════════════════════════
# 엑셀 출력
# ══════════════════════════════════════════════════════════════
COL_WIDTHS = {
    "카테고리": 16, "키워드": 18, "제목": 60, "언론사": 14,
    "발행시각": 17, "링크": 50, "네이버링크": 40, "요약초안": 55,
    "요약": 55, "메일카테고리": 14, "선택": 8,
}


def to_excel_bytes(df):
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="클리핑")
        ws = w.sheets["클리핑"]
        for i, col in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(col, 20)
        if "링크" in df.columns:
            link_col = list(df.columns).index("링크") + 1
            for row in range(2, len(df) + 2):
                c = ws.cell(row=row, column=link_col)
                if c.value:
                    c.hyperlink = c.value
                    c.style = "Hyperlink"
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
# 사이드바
# ══════════════════════════════════════════════════════════════
st.title("📰 상업용 부동산 뉴스 클리핑")
st.caption("네이버 뉴스 API 수집 · 본문 크롤링 + Gemini 명사형 요약")

with st.sidebar:
    st.header("⚙️ 설정")

    cid = get_secret("NAVER_CLIENT_ID")
    csecret = get_secret("NAVER_CLIENT_SECRET")
    if not (cid and csecret):
        cid = st.text_input("네이버 Client ID", type="password")
        csecret = st.text_input("네이버 Client Secret", type="password")
    else:
        st.caption("✓ 네이버 API 키 로드됨")

    hours = st.radio("수집 기간", [24, 48, 72],
                     format_func=lambda x: f"최근 {x}시간", index=0)

    st.divider()
    st.write("**관련성 필터**")
    use_relevance = st.checkbox("관련 없는 기사 걸러내기", value=True,
                                help="상업용 부동산·기관투자·부동산금융·정책과 "
                                     "무관한 기사를 목록에서 숨깁니다.")
    filter_level = st.selectbox("필터 강도", list(FILTER_LEVELS.keys()),
                                index=1, disabled=not use_relevance)
    rule_cutoff = FILTER_LEVELS[filter_level]

    st.divider()
    st.write("**중복 제거 민감도**")
    sim_threshold = st.slider("제목 유사도 임계값", 0.5, 0.9, 0.65, 0.05,
                              help="낮을수록 더 많이 제거")
    word_threshold = st.slider("단어 중복 임계값", 0.4, 0.9, 0.6, 0.05,
                               help="공통 단어 비율(자카드). 낮을수록 더 많이 제거")

    st.divider()
    st.write("**AI 요약 (Gemini)**")
    gemini_key = get_secret("GEMINI_API_KEY")
    model_override = ""
    if gemini_key:
        use_gemini = st.checkbox("Gemini로 요약", value=True)

        if st.button("🔌 모델 목록 새로고침 / 연결 테스트", **FULL_W):
            for k in ("_gemini_model_list", "_gemini_model"):
                st.session_state.pop(k, None)
            models, err = list_gemini_models(gemini_key)
            if models:
                st.session_state["_gemini_model_list"] = models
                st.success(f"✓ 연결 성공 · 후보 {len(models)}개")
            else:
                st.error(f"✗ {err}")

        cand, cerr = resolve_gemini_candidates(gemini_key)
        if cand:
            model_override = st.selectbox(
                "사용 모델", ["자동 (권장)"] + cand, index=0,
                help="자동 선택 시 사용 불가(404) 모델은 건너뛰고 다음 후보로 넘어갑니다.")
            model_override = "" if model_override == "자동 (권장)" else model_override
            st.caption(f"1순위 후보: `{cand[0]}`")
        elif cerr:
            st.caption(f"⚠️ {cerr}")
    else:
        use_gemini = False
        st.warning("⚠️ GEMINI_API_KEY 미설정\n\n"
                   "앱 Settings → Secrets에 추가하세요:\n\n"
                   '`GEMINI_API_KEY = "AIza..."`')

    st.divider()
    max_workers = st.slider("크롤링 동시 처리 수", 1, 8, 5,
                            help="높을수록 빠르지만 차단 위험 증가")

    st.divider()
    st.write("**카테고리 선택**")
    selected = {c: st.checkbox(c, value=True) for c in DEFAULT_KEYWORDS}


# ══════════════════════════════════════════════════════════════
# 키워드 편집 + 수집
# ══════════════════════════════════════════════════════════════
st.subheader("키워드 편집")
keyword_map = {}
cols = st.columns(2)
for i, (cat, kws) in enumerate(DEFAULT_KEYWORDS.items()):
    if selected.get(cat):
        with cols[i % 2]:
            txt = st.text_area(cat, value="\n".join(kws), height=110, key=f"kw_{cat}")
            keyword_map[cat] = [k.strip() for k in txt.splitlines() if k.strip()]

st.divider()

if st.button("🔍 뉴스 수집 시작", type="primary", **FULL_W):
    if not cid or not csecret:
        st.error("네이버 Client ID/Secret을 입력하세요.")
        st.stop()

    all_rows, errors, diags, kw_order = [], [], [], []
    total_kw = sum(len(v) for v in keyword_map.values())
    prog = st.progress(0.0, text="수집 중...")
    done = 0

    for cat, kws in keyword_map.items():
        for kw in kws:
            if kw not in kw_order:
                kw_order.append(kw)
            d = {"키워드": kw}
            rows, err = fetch_naver(kw, cat, cid, csecret, hours, diag=d)
            all_rows.extend(rows)
            diags.append(d)
            if err:
                errors.append(err)
            done += 1
            prog.progress(done / max(total_kw, 1), text=f"수집 중... ({kw})")
    prog.empty()

    if errors:
        st.error("네이버 API 오류:\n\n" + "\n\n".join(sorted(set(errors))))

    if diags:
        with st.expander("🔎 네이버 수집 진단", expanded=(len(all_rows) == 0)):
            dd = pd.DataFrame(diags)
            order = [c for c in ["키워드", "status", "raw_count", "kept", "newest"]
                     if c in dd.columns]
            st.dataframe(dd[order].rename(columns={
                "status": "HTTP상태", "raw_count": "원본건수",
                "kept": "기간내채택", "newest": "최신기사시각"}),
                hide_index=True, **FULL_W)

    if all_rows:
        dprog = st.progress(0.0, text="중복 제거 중...")
        df = dedup(pd.DataFrame(all_rows),
                   title_sim_threshold=sim_threshold,
                   word_sim_threshold=word_threshold,
                   progress_bar=dprog)
        dprog.empty()
    else:
        df = pd.DataFrame()

    # 재수집 시 하위 상태 초기화
    for k in ("editor_df", "mail_html", "collected", "editor", "restore_editor"):
        st.session_state.pop(k, None)

    if df.empty:
        st.warning("수집된 기사가 없습니다.")
    else:
        rank = {kw: i for i, kw in enumerate(kw_order)}
        df["_r"] = df["키워드"].map(rank).fillna(len(kw_order)).astype(int)
        df = df.sort_values(["_r", "발행시각"], ascending=[True, False])
        df = df.drop(columns="_r").reset_index(drop=True)

        # 관련성 점수 산출 + 규칙 기반 1차 제외
        df["관련도"] = [relevance_score(t, d) for t, d
                     in zip(df["제목"], df["요약초안"])]
        df["제외"] = (df["관련도"] < rule_cutoff) if use_relevance else False
        df["제외사유"] = ["규칙 필터" if x else "" for x in df["제외"]]

        st.session_state["collected"] = df
        st.session_state["collect_token"] = dt.datetime.now(KST).isoformat()

        kept_n = int((~df["제외"]).sum())
        cut_n = int(df["제외"].sum())
        st.success(f"✅ 총 {len(df)}건 (중복 제거 후 / 원본 {len(all_rows)}건)"
                   + (f" · 관련 {kept_n}건 / 제외 {cut_n}건" if cut_n else ""))
        st.dataframe(
            df[~df["제외"]]["키워드"].value_counts()
              .rename_axis("키워드").reset_index(name="건수"),
            hide_index=True)

        fname = f"뉴스클리핑_{dt.datetime.now(KST).strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            "📥 엑셀 다운로드", to_excel_bytes(df), file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            **FULL_W)


# ══════════════════════════════════════════════════════════════
# 배포 편집
# ══════════════════════════════════════════════════════════════
CATEGORY_RULES = [
    ("매입매각", ["매각", "매입", "매매", "인수", "거래", "딜 ", "클로징", "매각주관", "우선협상"]),
    ("개발계획", ["개발", "복합개발", "신축", "착공", "준공", "인허가", "부지", "설계", "기공"]),
    ("이전동향", ["이전", "사옥", "본사", "입주", "임차", "리모델링", "사무실"]),
    ("시장동향", ["시장", "전망", "공실", "임대료", "수익률", "지수", "동향", "거래량"]),
    ("정책", ["정책", "규제", "제도", "정부", "국토부", "세제", "금리", "개정", "법안"]),
    ("업계동향", ["운용사", "증권", "보험", "건설", "업계", "협회", "인사", "조직", "펀드", "리츠"]),
]


def suggest_category(keyword: str, title: str) -> str:
    """매칭 개수 기반 스코어링."""
    text = f"{keyword} {title}"
    best, best_score = "업계동향", 0
    for cat, words in CATEGORY_RULES:
        score = sum(2 if w in title else 1 for w in words if w in text)
        if score > best_score:
            best, best_score = cat, score
    return best


def build_mail_html(sel_df):
    FF = "'맑은 고딕', 'Malgun Gothic', sans-serif"

    def style(size, bold=False):
        return (f"line-height:1.8;font-family:{FF};color:#000000;orphans:2;"
                f"font-size:{size};font-weight:{'bold' if bold else '400'};"
                "margin:0px;padding:0px;")

    CATEGORY_STYLE = style("12pt", bold=True)
    BODY_STYLE = style("10pt")
    PRESS_STYLE = style("8pt")
    SPACER_STYLE = style("13pt")
    LINK_STYLE = (f"font-family:{FF};font-size:10pt;font-weight:bold;"
                  "color:#0000ff;text-decoration:underline;")
    BLANK = f'<p style="{SPACER_STYLE}">&nbsp;</p>'

    parts = [f'<div style="font-family:{FF};color:#000;">']
    for ci, cat in enumerate(MAIL_CATEGORIES):
        group = sel_df[sel_df["메일카테고리"] == cat]
        if group.empty:
            continue
        if ci > 0 and len(parts) > 1:
            parts.append(BLANK)
        parts.append(f'<p style="{CATEGORY_STYLE}">{html.escape(cat)}</p>')
        parts.append(BLANK)
        for _, row in group.iterrows():
            title = html.escape(str(row["제목"]))
            link = html.escape(str(row["링크"]), quote=True)
            summary = html.escape(str(row.get("요약", "") or ""))
            press = html.escape(str(row.get("언론사", "") or "").strip()) \
                or html.escape(PRESS_PLACEHOLDER)
            parts.append(
                f'<p style="{BODY_STYLE}"><a href="{link}" target="_blank" '
                f'rel="noopener noreferrer" style="{LINK_STYLE}">{title}</a></p>')
            for ln in summary.split("\n"):
                if ln.strip():
                    parts.append(f'<p style="{BODY_STYLE}">{ln.strip()}</p>')
            parts.append(f'<p style="{PRESS_STYLE}">{press}</p>')
            parts.append(BLANK)
    parts.append("</div>")
    return "".join(parts)


if "collected" in st.session_state and not st.session_state["collected"].empty:
    st.divider()
    st.header("✉️ 메일 배포용 정리")
    st.caption("배포할 기사를 선택하고 카테고리를 지정한 뒤 요약을 다듬으세요.")

    base = ensure_filter_cols(st.session_state["collected"].copy())
    token = st.session_state.get("collect_token", "")

    if st.session_state.get("editor_token") != token:
        edit = base.copy()
        edit.insert(0, "선택", False)
        edit["메일카테고리"] = edit.apply(
            lambda r: suggest_category(str(r.get("키워드", "")), str(r.get("제목", ""))), axis=1)
        edit["언론사"] = edit["언론사"].fillna("").apply(
            lambda s: s if str(s).strip() else PRESS_PLACEHOLDER)
        edit["요약"] = edit["요약초안"].fillna("").apply(lambda t: first_sentence(t, 70))
        st.session_state["editor_df"] = ensure_filter_cols(edit)
        st.session_state["editor_token"] = token

    if st.session_state.pop("_flash", None):
        st.success(st.session_state.pop("_flash_msg", "완료"))

    full_df = ensure_filter_cols(st.session_state["editor_df"].copy())
    st.session_state["editor_df"] = full_df
    visible_df = full_df[~full_df["제외"]]
    hidden_df = full_df[full_df["제외"]]

    # ── AI 관련성 재판정 ─────────────────────────────────────
    fc1, fc2 = st.columns([2, 1])
    ambiguous = visible_df[visible_df["관련도"] < AI_CHECK_BELOW]
    with fc1:
        st.caption(f"표시 {len(visible_df)}건 · 규칙으로 제외 {len(hidden_df)}건 · "
                   f"AI 재판정 대상(관련도 낮음) {len(ambiguous)}건")
    with fc2:
        run_ai_filter = st.button(
            f"🤖 AI로 관련성 재판정 ({len(ambiguous)}건)", **FULL_W,
            disabled=ambiguous.empty or not (use_gemini and gemini_key),
            help="제목만 40개씩 묶어 보내므로 호출 수가 적습니다.")

    if run_ai_filter:
        cands, _e = ([model_override], None) if model_override \
            else resolve_gemini_candidates(gemini_key)
        if not cands:
            st.error("Gemini 모델을 확인할 수 없습니다.")
        else:
            picker_f = ModelPicker(cands)
            titles = ambiguous["제목"].astype(str).tolist()
            with st.spinner(f"제목 {len(titles)}건 관련성 판정 중..."):
                keep_pos, ferrs = ai_relevance_check(titles, gemini_key, picker_f)
            drop_idx = [ix for n, ix in enumerate(ambiguous.index) if n not in keep_pos]
            new_df = full_df.copy()
            for ix in drop_idx:
                new_df.loc[ix, "제외"] = True
                new_df.loc[ix, "제외사유"] = "AI 판정"
            st.session_state["editor_df"] = new_df
            st.session_state.pop("editor", None)
            st.session_state["_flash"] = True
            st.session_state["_flash_msg"] = (
                f"✅ AI 판정 완료 · {len(drop_idx)}건 추가 제외"
                + (f" (오류 {len(ferrs)}건은 유지)" if ferrs else ""))
            st.rerun()

    # ── 제외된 기사 확인 · 복원 ───────────────────────────────
    if not hidden_df.empty:
        with st.expander(f"🚫 제외된 기사 {len(hidden_df)}건 (확인 · 복원)"):
            restore_src = hidden_df.copy()
            restore_src.insert(0, "복원", False)
            restored = st.data_editor(
                restore_src, hide_index=True, **FULL_W, height=260,
                column_order=["복원", "관련도", "제외사유", "제목", "언론사", "링크"],
                column_config={
                    "복원": st.column_config.CheckboxColumn("복원", width="small"),
                    "관련도": st.column_config.NumberColumn("점수", width="small"),
                    "제외사유": st.column_config.TextColumn("사유", width="small"),
                    "제목": st.column_config.TextColumn("제목", width=col_width(520)),
                    "언론사": st.column_config.TextColumn("언론사", width="small"),
                    "링크": st.column_config.LinkColumn("링크", display_text="열기",
                                                      width=col_width(70, "small")),
                },
                disabled=["관련도", "제외사유", "제목", "언론사", "링크"],
                key="restore_editor",
            )
            rc1, rc2 = st.columns(2)
            if rc1.button("↩️ 선택한 기사 복원", **FULL_W):
                back_idx = restored[restored["복원"] == True].index
                nd = full_df.copy()
                for ix in back_idx:
                    nd.loc[ix, "제외"] = False
                    nd.loc[ix, "제외사유"] = ""
                st.session_state["editor_df"] = nd
                for k in ("editor", "restore_editor"):
                    st.session_state.pop(k, None)
                st.session_state["_flash"] = True
                st.session_state["_flash_msg"] = f"✅ {len(back_idx)}건 복원"
                st.rerun()
            if rc2.button("↩️ 전체 복원 (필터 해제)", **FULL_W):
                nd = full_df.copy()
                nd["제외"] = False
                nd["제외사유"] = ""
                st.session_state["editor_df"] = nd
                for k in ("editor", "restore_editor"):
                    st.session_state.pop(k, None)
                st.session_state["_flash"] = True
                st.session_state["_flash_msg"] = "✅ 전체 복원"
                st.rerun()

    edited_df = st.data_editor(
        visible_df,
        hide_index=True, **FULL_W, height=460,
        column_order=["선택", "키워드", "제목", "요약",
                      "언론사", "발행시각", "링크"],
        column_config={
            "선택": st.column_config.CheckboxColumn("선택", width="small"),
            "키워드": st.column_config.TextColumn("키워드", width=col_width(110, "small")),
            "메일카테고리": None,   # 분류는 생성 시 AI가 결정
            "관련도": None, "제외": None, "제외사유": None,
            "제목": st.column_config.TextColumn("제목", width=col_width(560)),
            "요약": st.column_config.TextColumn("요약 (직접 수정)", width=col_width(400)),
            "언론사": st.column_config.TextColumn("언론사 (직접 수정)", width=col_width(120, "small")),
            "발행시각": st.column_config.TextColumn("발행시각", width=col_width(125, "small")),
            "링크": st.column_config.LinkColumn("링크", display_text="열기",
                                              width=col_width(70, "small")),
            "요약초안": None, "카테고리": None, "네이버링크": None,
        },
        disabled=["제목", "키워드", "발행시각", "링크"],
        key="editor",
    )

    sel = edited_df[edited_df["선택"] == True].copy()   # 원본 인덱스 유지
    st.write(f"선택된 기사: **{len(sel)}건**")

    if not sel.empty:
        need_press = sel[sel["언론사"].astype(str).str.strip().isin(["", PRESS_PLACEHOLDER])]
        if not need_press.empty:
            st.caption(f"ℹ️ 선택한 기사 중 {len(need_press)}건은 언론사가 비어 있습니다. "
                       "메일 본문 생성 시 자동으로 채워집니다.")

    make_mail = st.button("📋 메일 본문 생성", type="primary",
                          **FULL_W, disabled=sel.empty)

    if make_mail:
        use_ai = bool(use_gemini and gemini_key)
        picker = None
        if use_ai:
            if model_override:
                cand_list, cerr2 = [model_override], None
            else:
                cand_list, cerr2 = resolve_gemini_candidates(gemini_key)
            if not cand_list:
                st.error(f"Gemini 모델 확인 실패 → 첫 문장 요약으로 대체합니다. ({cerr2})")
                use_ai = False
            else:
                # 오늘 이미 한도가 소진된 모델은 다시 시도하지 않는다 (쿼터 절약)
                spent = st.session_state.get("_gemini_spent", {})
                alive = [m for m in cand_list if m not in spent] or cand_list
                picker = ModelPicker(alive)
                st.info(f"Gemini 1순위 모델: `{alive[0]}`"
                        + (f" (실패 시 {len(alive) - 1}개 후보로 자동 폴백)"
                           if len(alive) > 1 else "")
                        + (f" · 한도 소진으로 제외한 모델 {len(spent)}개" if spent else ""))

        sel_copy = sel.copy()   # 원본 인덱스 유지 → 편집표에 되돌려쓰기 가능
        prog = st.progress(0.0, text="본문 크롤링 및 요약 생성 중...")
        logs, ok, press_filled, title_fixed, cat_ai, naver_fb, desc_ai = [], 0, 0, 0, 0, 0, 0
        total_n = len(sel_copy)

        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = [
                ex.submit(summarize_one, i, str(r.get("링크", "")),
                          str(r.get("제목", "")), str(r.get("요약초안", "")),
                          gemini_key, picker, use_ai)
                for i, r in sel_copy.iterrows() if str(r.get("링크", ""))
            ]
            for n, fut in enumerate(as_completed(futures), start=1):
                try:
                    idx, category, summary, press, page_title, log = fut.result()
                    if summary:
                        sel_copy.loc[idx, "요약"] = summary
                        if log is None:
                            ok += 1
                    # AI 분류 우선 적용 (실패 시 규칙 기반 추천값 유지)
                    if category in MAIL_CATEGORIES:
                        if sel_copy.loc[idx, "메일카테고리"] != category:
                            cat_ai += 1
                        sel_copy.loc[idx, "메일카테고리"] = category
                    cur = str(sel_copy.loc[idx, "언론사"]).strip()
                    if press and cur in ("", PRESS_PLACEHOLDER, "nan"):
                        sel_copy.loc[idx, "언론사"] = press
                        press_filled += 1
                    # 네이버 제목이 잘려 있으면 기사 원문 제목으로 교체
                    cur_title = str(sel_copy.loc[idx, "제목"])
                    fixed = better_title(cur_title, page_title)
                    if fixed != cur_title:
                        sel_copy.loc[idx, "제목"] = fixed
                        title_fixed += 1
                    # 크롤링이 막힌 매체 → 네이버 검색 요약문으로 대체 (빈칸 방지)
                    if not str(sel_copy.loc[idx, "요약"] or "").strip():
                        fb = naver_fallback_summary(sel_copy.loc[idx, "요약초안"])
                        if fb:
                            sel_copy.loc[idx, "요약"] = fb
                            naver_fb += 1
                            if log:
                                log += " → 네이버 요약문으로 대체"
                    if log:
                        if "발췌문으로 AI 요약" in log:
                            desc_ai += 1
                        logs.append(log)
                except Exception as e:
                    logs.append(f"워커 오류: {str(e)[:100]}")
                prog.progress(n / max(total_n, 1), text=f"처리 중... ({n}/{total_n})")
        prog.empty()

        # 한도 소진 모델을 세션에 기억해 다음 실행에서 건너뛴다
        if picker is not None and picker.dead:
            spent = dict(st.session_state.get("_gemini_spent", {}))
            spent.update({m: w for m, w in picker.dead.items() if "일일 한도" in w})
            st.session_state["_gemini_spent"] = spent
            if not picker.current():
                st.error(
                    "🚫 무료 Gemini 일일 사용량을 모두 소진했습니다. 요약은 기사 첫 문장으로 "
                    "대체됐습니다.\n\n"
                    "· 한도 초기화: 태평양 시간 자정 = **한국시간 다음날 오후 4시**\n"
                    "· 지금 바로 쓰려면 Google AI Studio에서 결제 연결(유료 티어) 필요")

        label = "Gemini 요약" if use_ai else "첫 문장 요약"
        st.write(f"**결과:** ✓ {ok}/{total_n}건 {label} 성공"
                 + (f" · 분류 {cat_ai}건 AI 재조정" if cat_ai else "")
                 + (f" · 언론사 {press_filled}건 보완" if press_filled else "")
                 + (f" · 잘린 제목 {title_fixed}건 복원" if title_fixed else "")
                 + (f" · 본문 차단 {desc_ai}건은 발췌문으로 AI 요약" if desc_ai else "")
                 + (f" · {naver_fb}건은 발췌문 원문 사용" if naver_fb else ""))

        still_empty = sel_copy[sel_copy["언론사"].astype(str).str.strip()
                               .isin(["", PRESS_PLACEHOLDER, "nan"])]
        if not still_empty.empty:
            st.warning(f"⚠️ 언론사 미확인 {len(still_empty)}건 — 위 표에서 직접 입력하세요.")

        if logs:
            st.warning(f"⚠️ {len(logs)}건 문제 발생 (첫 문장으로 대체됨)")
            with st.expander("🔍 실패 원인 상세"):
                for log in logs[:20]:
                    st.text(f"• {log}")

        ordered = sel_copy.copy()
        ordered["_c"] = ordered["메일카테고리"].map(
            {c: i for i, c in enumerate(MAIL_CATEGORIES)})
        ordered = ordered.sort_values(["_c", "발행시각"], ascending=[True, False])
        st.session_state["mail_html"] = build_mail_html(ordered)

        # 생성 결과를 편집표에도 반영
        back = st.session_state["editor_df"].copy()
        for idx in sel_copy.index:
            for col in ("요약", "언론사", "제목", "메일카테고리"):
                back.loc[idx, col] = sel_copy.loc[idx, col]
        st.session_state["editor_df"] = back

        st.success("✅ 메일 본문이 생성되었습니다.")

    if "mail_html" in st.session_state:
        st.subheader("메일 본문")
        st.caption("[메일 본문 복사] 버튼을 누르면 서식이 클립보드에 담깁니다.")
        mail_html = st.session_state["mail_html"]

        import json as _json
        html_js = _json.dumps(mail_html)
        copy_widget = f"""
        <div id="src" style="position:absolute;left:-9999px;top:-9999px;">{mail_html}</div>
        <div style="font-family:'맑은 고딕',sans-serif;">
          <button id="copyBtn" style="padding:10px 18px;font-size:14px;
              background:#ff4b4b;color:#fff;border:none;border-radius:6px;
              cursor:pointer;width:100%;font-weight:bold;">
            📋 메일 본문 복사 (서식 유지)
          </button>
          <span id="copyMsg" style="margin-left:10px;color:#0a0;font-weight:bold;"></span>
        </div>
        <script>
        const htmlStr = {html_js};
        document.getElementById('copyBtn').addEventListener('click', async () => {{
          const src = document.getElementById('src');
          try {{
            await navigator.clipboard.write([new ClipboardItem({{
              'text/html': new Blob([htmlStr], {{type: 'text/html'}}),
              'text/plain': new Blob([src.innerText], {{type: 'text/plain'}})
            }})]);
            document.getElementById('copyMsg').innerText = '✓ 복사됨! 메일에 붙여넣으세요';
          }} catch (e) {{
            const range = document.createRange();
            range.selectNode(src);
            window.getSelection().removeAllRanges();
            window.getSelection().addRange(range);
            document.execCommand('copy');
            window.getSelection().removeAllRanges();
            document.getElementById('copyMsg').innerText = '✓ 복사됨 (폴백)';
          }}
        }});
        </script>
        """
        st.components.v1.html(copy_widget, height=100, scrolling=False)

        with st.expander("📨 메일 본문 미리보기", expanded=True):
            st.components.v1.html(
                f"<div style=\"font-family:'맑은 고딕',sans-serif;\">{mail_html}</div>",
                height=500, scrolling=True)

        full_html = ("<!doctype html><html><head><meta charset='utf-8'></head><body>"
                     + mail_html + "</body></html>")
        st.download_button(
            "📥 (백업) 메일 HTML 파일 다운로드", data=full_html.encode("utf-8"),
            file_name=f"뉴스클리핑_메일_{dt.datetime.now(KST).strftime('%Y%m%d')}.html",
            mime="text/html", **FULL_W)
